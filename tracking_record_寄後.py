from IPython import get_ipython
get_ipython().magic('reset -f')
import pandas as pd
import pygsheets 
import pyodbc
import concurrent.futures
import numpy as np
import openpyxl
from datetime import datetime, timedelta
from sqlalchemy import create_engine

# =============================================================================
# # Load the existing Excel file
# workbook = openpyxl.load_workbook('trackingrecord_寄後.xlsx')
# gc = pygsheets.authorize(service_file='C:/Users/11020984/Desktop/tracking_record_insert/tracking-record-time-7c7ffa4f8b1a.json')
# # Define the date range
# end_date = datetime.now()
# start_date = end_date - timedelta(days=3)  # Change the number of days as needed
# # Create a date format for the sheet names
# date_format = '%m%d'
# #existing_excel_file = 'path_to_your_existing_excel.xlsx'
# # Get all the sheet names within the date range
# date_range = [(end_date - timedelta(days=i)).strftime(date_format) for i in range((end_date - start_date).days, 0, -1)]
# #workbook = openpyxl.load_workbook('tracking_record.xlsx')
# dataframes = []
# selected_columns = ['邀約日期', '資料區域群組', '預計發送對象聯絡人代號', '預計發送對象姓名', '公司名稱', '公司代號', '電訪人員', '(1)未接1\n(2)未接2\n(3)資料已確認\n(4)接通掛電話\n(5)資料無效',
#                     '最多打兩通\n(1)有收到\n(2)沒收到1、2\n(3)不清楚1、2',
#                     '(1)可線上介紹\n(2)同意約訪(業務拜訪)\n(3)有需要再約\n(4)完全沒興趣',
#                     '有需要再約二次約訪\n(1)可線上介紹\n(2)有需要再約', '其他備註', '業務安排K大/拜訪日\n(可確定行程再補填)',
#                     '加入LINE好友\n(1)有\n(2)無', '補寄地址(沒收到)', 'LINEID', '聯繫方式\n(1)電訪\n(2)通訊軟體',
#                     '是否離職']
#                     
# 
# # Iterate through the sheet names and import data
# for sheet_name in date_range:
#     try:
#     # Fetch data from Google Sheet
#         spreadsheet = gc.open_by_key('1w6c60VixiIqx_tGlYUSF8hW650h7sl9kUrsvYNHlS38')
#         worksheet = spreadsheet.worksheet_by_title(sheet_name)
#         data = worksheet.get_all_values()
#         # Convert data to DataFrame
#         #df = pd.DataFrame(data[1:], columns=data[0])
#         df = pd.DataFrame(data[1:], columns=data[0])
#         # Reset index to ensure it's unique
#         df.reset_index(drop=True, inplace=True)
#         df.replace('', np.nan, inplace=True)
#         current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#         #df.dropna(subset=['聯絡人代號'], inplace=True)
#         if sheet_name in workbook.sheetnames:
#            continue
#         else:
#            worksheet = workbook.create_sheet(sheet_name)
#            for row_data in data:
#                row_data.append(current_time)
#                worksheet.append(row_data)
#            #dataframes.append(df)
#            dataframes.append(df.loc[:, selected_columns])
#     except pygsheets.exceptions.WorksheetNotFound:
#        # Handle the case where the worksheet is not found
#        pass
# # Concatenate the selected DataFrames vertically
# oldsheet_df = pd.concat(dataframes, axis=0, ignore_index=True)
# #print(oldsheet_df.columns)
# # Save the changes to the existing Excel file
# workbook.save('trackingrecord_寄後.xlsx')
# # Close the workbook
# workbook.close()
# '''
# drop nan rows and no values in '是否接通', '是否同意邀約', '是否離職', '備註'
# '''
# oldsheet_df.dropna(how='all', inplace=True)
# oldsheet_df.dropna(subset=['(1)未接1\n(2)未接2\n(3)資料已確認\n(4)接通掛電話\n(5)資料無效',
#                     '最多打兩通\n(1)有收到\n(2)沒收到1、2\n(3)不清楚1、2',
#                     '(1)可線上介紹\n(2)同意約訪(業務拜訪)\n(3)有需要再約\n(4)完全沒興趣',
#                     '有需要再約二次約訪\n(1)可線上介紹\n(2)有需要再約', '其他備註', '業務安排K大/拜訪日\n(可確定行程再補填)',
#                     '加入LINE好友\n(1)有\n(2)無', '補寄地址(沒收到)', 'LINEID', '聯繫方式\n(1)電訪\n(2)通訊軟體',
#                     '是否離職'], how='all', inplace=True)
# =============================================================================
end_date = datetime.now()
start_date = end_date - timedelta(days=3)  # Change the number of days as needed
date_format = '%Y-%m-%d'

date_range = [(end_date - timedelta(days=i)).strftime(date_format) for i in range((end_date - start_date).days, 0, -1)]
dataframes = []
selected_columns = ['邀約日期', '資料區域群組', '預計發送對象聯絡人代號', '預計發送對象姓名', '公司名稱', '公司代號', '電訪人員', '(1)未接1\n(2)未接2\n(3)資料已確認\n(4)接通掛電話\n(5)資料無效',
                    '最多打兩通\n(1)有收到\n(2)沒收到1、2\n(3)不清楚1、2',
                    '(1)可線上介紹\n(2)同意約訪\n(3)有需要再約\n(4)完全沒興趣',
                    '有需要再約二次約訪\n(1)可線上介紹\n(2)有需要再約', '其他備註', '業務安排K大/拜訪日\n(可確定行程再補填)',
                    '加入LINE好友\n(1)有\n(2)無', '補寄地址(沒收到)', 'LINEID', '聯繫方式\n(1)電訪\n(2)通訊軟體',
                    '是否離職']
time_df = pd.read_excel(r'Z:\02_台灣事業部\1.北區\13.業務管理組\CRM共用資料夾\業助工作資料\業助資料夾\●皓皓●\4.K大預約總表\K大邀約及執行統計表\如苾\廣發型錄寄後電訪.xlsx', 
                            usecols=selected_columns)
#sheet_name = '2023-09-25'
#oldsheet_df = time_df[time_df['邀約日期']==sheet_name]
oldsheet_df = time_df[time_df['邀約日期'] == '10-12']
oldsheet_df.reset_index(drop=True, inplace=True)
#oldsheet_df.reset_index(drop=True, inplace=True)

'''
connect to CRM
'''
server = '192.168.1.112'
#database = 'TEMP_GDCRM'
database = 'GDCRM'
username = 'sa'
password = 'mis528716#'
# Establish a connection
crm_con = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password) 
cursor = crm_con.cursor()
# Create SQLAlchemy engine to connect to SQL Server
engine = create_engine('mssql+pyodbc:///?odbc_connect={}'.format('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password))
'''
assign columns
'''
oldsheet_df['CreateDate'] = datetime.now().replace(microsecond=0) - timedelta(hours=8)
oldsheet_df['ModifyDate'] = oldsheet_df['CreateDate']
oldsheet_df.rename(columns={"預計發送對象聯絡人代號": "ContactID",
                            "(1)未接1\n(2)未接2\n(3)資料已確認\n(4)接通掛電話\n(5)資料無效": "狀態",
                            "最多打兩通\n(1)有收到\n(2)沒收到1、2\n(3)不清楚1、2": "型錄是否收到",
                            "(1)可線上介紹\n(2)同意約訪\n(3)有需要再約\n(4)完全沒興趣":"是否同意約訪",
                            "有需要再約二次約訪\n(1)可線上介紹\n(2)有需要再約":"是否需要二次約訪",
                            "業務安排K大/拜訪日\n(可確定行程再補填)":"安排的拜訪日",
                            "加入LINE好友\n(1)有\n(2)無":"是否加入LINE好友",
                            "補寄地址(沒收到)":"補寄型錄地址",
                            "聯繫方式\n(1)電訪\n(2)通訊軟體":"聯繫方式",
                            "公司代號": "CompanyID","預計發送對象姓名": "ContactName",
                            "邀約日期": "RecordDate"}, inplace=True)
oldsheet_df['DeptID'] = '003'
oldsheet_df['CreateDeptID'] = '003'
oldsheet_df['OwnerDeptID'] = '003'
oldsheet_df['ModifyDeptID'] = '003'
oldsheet_df['CreatorID'] = '11020984'
oldsheet_df['WorkProperty'] = '0'
oldsheet_df['VisitHours'] = '0'
oldsheet_df['VisitArea'] = '無'
budget_item_columns = ['BudgetItem1', 'BudgetItem2', 'BudgetItem3', 'BudgetItem4', 'BudgetItem5', 'BudgetItem6', 'TotalBudget']
expense_others_columns = ['ExpenseItem1', 'ExpenseItem2', 'ExpenseItem3', 'ExpenseItem4', 'ExpenseItem5', 'ExpenseItem5', 'TotalExpense',
                          'IsTask', 'IsIntroduce','UpdateVisitAppointment','UpdatePhoneAppointment', 'chkflag','WorkHours']
# Initialize budget and expense item columns with 0
oldsheet_df[budget_item_columns + expense_others_columns] = 0
oldsheet_df.fillna(' ', inplace=True)
oldsheet_df['ContactID']= oldsheet_df['ContactID'].astype(str).str.zfill(8)
oldsheet_df['CompanyID']= oldsheet_df['CompanyID'].astype(str).str.zfill(8)
# Get the current year
current_year = datetime.now().year
# Function to update RecordDate
def update_record_date(record_date):
    return f"{current_year}-{record_date[:2]}{record_date[2:]} 00:00:00"
# Apply the function to update RecordDate
oldsheet_df["RecordDate"] = oldsheet_df["RecordDate"].apply(
    lambda x: update_record_date(x) if x else ''
)

def create_work_description(row):
    work_description = []
    columns_to_include = ['狀態',
    '型錄是否收到',
    '是否同意約訪',
    '是否需要二次約訪', '其他備註', '安排的拜訪日',
    '是否加入LINE好友', '補寄型錄地址', 'LINEID', '聯繫方式',
    '是否離職']
    for column in columns_to_include:
        value = row[column]
        if value:
            work_description.append(f"{column}: {value}")
    return '\n\n'.join(work_description)
oldsheet_df['WorkDescription'] = oldsheet_df.apply(create_work_description, axis=1)
'''
select ContactID for ContactName
select EmplID for SalesID
'''
sales_df_names = ', '.join([f"N'{name}'" for name in list(oldsheet_df['電訪人員'].unique())])
contact_info = [
    (contact_id, contactName, company_id)
    for contact_id, contactName, company_id in oldsheet_df[['ContactID', 'ContactName', 'CompanyID']].drop_duplicates().values
    if pd.isna(contact_id)
]
has_nan_contact_id = len(contact_info) > 0
if has_nan_contact_id:
    #contact_df_names = ', '.join([f"N'{name}'" for name, _ in contact_info])
    contact_company_ids = ', '.join([f"'{company_id}'" for _,_, company_id in contact_info])
    contact_names = ', '.join([f"'{contactName}'" for _,contactName,_ in contact_info])
    query_id = f'''SELECT EmplID AS 'SalesID', DisplayName AS '電訪人員' 
                       FROM [GDCRM].[dbo].[Employee] WHERE DisplayName IN ({sales_df_names})'''
    query_id2 = f'''SELECT ContactID, OrigID AS 'CompanyID' 
                       FROM [GDCRM].[dbo].[Contact] 
                       WHERE DisplayName IN ({contact_names}) AND OrigID IN ({contact_company_ids})'''
    id_df1 = pd.read_sql_query(query_id, crm_con)
    id_df2 = pd.read_sql_query(query_id2, crm_con)
    oldsheet_df = oldsheet_df.merge(id_df1, on='電訪人員', how='left')
    oldsheet_df = oldsheet_df.merge(id_df2, on='CompanyID', how='left')
    oldsheet_df['ContactID'] = oldsheet_df['ContactID_x'].combine_first(oldsheet_df['ContactID_y'])
    oldsheet_df = oldsheet_df.drop(['ContactID_x', 'ContactID_y'], axis=1)
else: 
    query_id = f'''
    SELECT EmplID AS 'SalesID', DisplayName AS '電訪人員' 
    FROM [GDCRM].[dbo].[Employee] 
    WHERE DisplayName IN ({sales_df_names});
    '''                                                                                                                                                            
    id_df = pd.read_sql_query(query_id, crm_con)
    oldsheet_df = oldsheet_df.merge(id_df, on='電訪人員', how='left')
oldsheet_df['OwnerID'] = oldsheet_df['SalesID']
oldsheet_df['ModifierID'] = oldsheet_df['SalesID']

if oldsheet_df['SalesID'].isna().any():
   oldsheet_df.loc[oldsheet_df['SalesID'].isna(), 'SalesID'] = 'A13066'
   oldsheet_df['WorkDescription'] = oldsheet_df.apply(
        lambda row: row['WorkDescription'] + '\n\n電訪人員: ' + row['電訪人員'] if row['SalesID'] == 'A13066' else row['WorkDescription'],
        axis=1
    )
oldsheet_df['OwnerID'] = oldsheet_df['SalesID']
oldsheet_df['ModifierID'] = oldsheet_df['SalesID']
'''
assign SalesWorkTypeID and WorkHours
'''
def map_work_type(row):
    if (row['狀態'] == '資料已確認' and row['聯繫方式'] =='通訊軟體'):
        return 'B7-1'
    elif row['狀態']== '資料無效' or row['狀態']== '接通掛電話':
        return 'C12-2'
    elif row['狀態']== '未接2' :
        return 'B7-2' 
    elif row['狀態']== '未接1' :
        return 'B2'
    else:
        return 'B7'
oldsheet_df['SalesWorkTypeID'] = oldsheet_df.apply(map_work_type, axis=1)
'''
update contact table if we have these two conditions (停用 and 離職)
'''
keywords = ['退休', '過世', '轉行', '空號', '歿', '沒有在做']
filtered_rows_stop = oldsheet_df[
    oldsheet_df["狀態"] .str.contains("資料無效")| oldsheet_df['其他備註'].str.contains('|'.join(keywords))]
for index, row in filtered_rows_stop.iterrows():
    stop_query = f"""UPDATE Contact
    SET DisplayName = CONCAT(DisplayName, '(停用)'), Disable = '1'
    WHERE ContactID = {row['ContactID']}
    """
    cursor.execute(stop_query)
    crm_con.commit()

#quit the job
keywords = ['轉行', '過世', '產業', '退休', '沒有在做', '已約']
filtered_rows_leave = oldsheet_df[
    (oldsheet_df["是否離職"] == "是") & ~oldsheet_df['狀態'].str.contains("資料無效") 
    & ~oldsheet_df['其他備註'].str.contains('|'.join(keywords)) | (oldsheet_df['其他備註'].str.contains('離職'))]

for index, row in filtered_rows_leave.iterrows():
    leave_query = f"""UPDATE Contact
    SET DisplayName = CONCAT(DisplayName, '(離職)'), IsLeave = '1'
    WHERE ContactID = {row['ContactID']}
    """
    cursor.execute(leave_query)
    crm_con.commit()

'''
lineId
'''
filtered_rows_line = oldsheet_df[oldsheet_df['LINEID'].str.contains(r'^[a-zA-Z0-9]+$')]
print(filtered_rows_line.columns)
for index, row in filtered_rows_line.iterrows():
    leave_query = f"""UPDATE Contact
    SET ICQ = '{row['LINEID']}'
    WHERE ContactID = {row['ContactID']}
    """
    cursor.execute(leave_query)
    crm_con.commit()
    


'''
drop unnessary columns
'''  
oldsheet_df = oldsheet_df.drop(columns=['資料區域群組', '電訪人員', '公司名稱', '狀態',
                                        '型錄是否收到',
                                        '是否同意約訪',
                                        '是否需要二次約訪', '其他備註', '安排的拜訪日',
                                        '是否加入LINE好友', '補寄型錄地址', 'LINEID', '聯繫方式',
                                        '是否離職'])
'''
find the max_id from KeyFieldDefine and update table
'''
#Function to fetch the current Seed value from KeyFieldDefine
def fetch_seed_value():
    query_maxid = "SELECT Seed FROM KeyFieldDefine WHERE TableName = 'TrackingRecord'"
    cursor.execute(query_maxid)
    seed_max = cursor.fetchone()[0]
    return seed_max
#Function to update the Seed value in the KeyFieldDefine table
def update_seed_value(seed_max, oldsheet_df_len):
    max_id = seed_max + oldsheet_df_len
    update_seed = f"UPDATE KeyFieldDefine SET Seed = {max_id}, chkflag = {max_id} WHERE TableName = 'TrackingRecord'"
    cursor.execute(update_seed)
    crm_con.commit()
#Execute the fetch_seed_value function asynchronously
with concurrent.futures.ThreadPoolExecutor() as executor:
    future_seed = executor.submit(fetch_seed_value)
#Wait for the fetch_seed_value function to complete and get the Seed value
seed_max = future_seed.result()
#Call the update_seed_value function with the fetched Seed value and the length of oldsheet_df
update_seed_value(seed_max, len(oldsheet_df.index))
#Calculate the updated max_id
max_id = seed_max + len(oldsheet_df.index)

#Now 'max_id' holds the updated value

'''
assign TrackingRecordID
'''
# Calculate the range of values to add
start_value = max_id - len(oldsheet_df.index) + 1
end_value = max_id + 1
# Create a new 'TaskID' column with incremented values
oldsheet_df['TrackingRecordID'] = [f'{x:08}' for x in range(start_value, end_value)]
oldsheet_df['WorkDescription'] = oldsheet_df.apply(
        lambda row: row['WorkDescription'] + '\n\nTrackingRecordID: ' + row['TrackingRecordID'],
        axis=1
    )

'''
insert Trackingrecord
'''
oldsheet_df.to_sql('TrackingRecord', engine, if_exists='append', index=False)    
 
crm_con.commit()
cursor.close()
crm_con.close()